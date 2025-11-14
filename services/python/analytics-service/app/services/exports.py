"""Export service for data export functionality."""

import csv
import os
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from clickhouse_connect.driver import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core import redis_manager, settings
from app.core.logging import get_logger
from app.schemas.exports import ExportFormat, ExportRequest, ExportResponse, ExportStatus

logger = get_logger(__name__)


class ExportService:
    """Service for handling data exports."""

    def __init__(self, clickhouse: Client):
        """Initialize service with ClickHouse client."""
        self.ch = clickhouse
        self.export_dir = settings.EXPORT_TEMP_DIR

    async def process_export(self, export_id: str, request: ExportRequest) -> None:
        """Process export request asynchronously."""
        try:
            # Update status to processing
            await self._update_export_status(export_id, ExportStatus.PROCESSING)
            
            # Execute query
            result = self.ch.query(request.query, parameters=request.query_params or {})
            
            # Check row limit
            if result.row_count > request.row_limit:
                await self._update_export_status(
                    export_id, 
                    ExportStatus.FAILED, 
                    error=f"Result exceeds row limit of {request.row_limit}"
                )
                return
            
            # Generate export file
            file_path = await self._generate_export_file(
                export_id,
                result,
                request.format,
                request.filename or f"export_{export_id}"
            )
            
            # Update status to completed
            await self._update_export_status(
                export_id,
                ExportStatus.COMPLETED,
                file_path=file_path,
                row_count=result.row_count
            )
            
            logger.info("export_completed", export_id=export_id, rows=result.row_count)
            
        except Exception as e:
            logger.error("export_processing_failed", export_id=export_id, error=str(e))
            await self._update_export_status(export_id, ExportStatus.FAILED, error=str(e))

    async def _generate_export_file(
        self,
        export_id: str,
        result: Any,
        format: ExportFormat,
        filename: str
    ) -> str:
        """Generate export file in specified format."""
        os.makedirs(self.export_dir, exist_ok=True)
        
        if format == ExportFormat.CSV:
            file_path = os.path.join(self.export_dir, f"{filename}.csv")
            self._write_csv(result, file_path)
        else:  # EXCEL
            file_path = os.path.join(self.export_dir, f"{filename}.xlsx")
            self._write_excel(result, file_path)
        
        return file_path

    def _write_csv(self, result: Any, file_path: str) -> None:
        """Write result to CSV file."""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(result.column_names)
            writer.writerows(result.result_rows)

    def _write_excel(self, result: Any, file_path: str) -> None:
        """Write result to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write header
        for col_idx, col_name in enumerate(result.column_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row_data in enumerate(result.result_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(file_path)

    async def _update_export_status(
        self,
        export_id: str,
        status: ExportStatus,
        file_path: Optional[str] = None,
        row_count: Optional[int] = None,
        error: Optional[str] = None
    ) -> None:
        """Update export status in Redis."""
        export_data = {
            "export_id": export_id,
            "status": status,
            "updated_at": datetime.utcnow().isoformat(),
        }
        
        if file_path:
            export_data["file_path"] = file_path
        if row_count is not None:
            export_data["row_count"] = row_count
        if error:
            export_data["error"] = error
        
        await redis_manager.setex(
            f"export:{export_id}",
            settings.EXPORT_RETENTION_HOURS * 3600,
            export_data
        )

    async def get_export_status(self, export_id: str) -> Optional[ExportResponse]:
        """Get export status from Redis."""
        data = await redis_manager.get(f"export:{export_id}")
        if not data:
            return None
        
        return ExportResponse(**data)

    async def get_export_file(self, export_id: str) -> Optional[str]:
        """Get export file path."""
        data = await redis_manager.get(f"export:{export_id}")
        if not data or data.get("status") != ExportStatus.COMPLETED:
            return None
        
        return data.get("file_path")

    async def delete_export(self, export_id: str) -> bool:
        """Delete export file and metadata."""
        data = await redis_manager.get(f"export:{export_id}")
        if not data:
            return False
        
        # Delete file
        if "file_path" in data and os.path.exists(data["file_path"]):
            os.remove(data["file_path"])
        
        # Delete Redis key
        await redis_manager.delete(f"export:{export_id}")
        
        return True

    async def execute_query(self, query: str) -> Any:
        """Execute a query and return result."""
        return self.ch.query(query)

    async def generate_metrics_excel(
        self,
        start_date: datetime,
        end_date: datetime,
        file_path: str
    ) -> None:
        """Generate comprehensive metrics Excel workbook."""
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # User metrics sheet
        await self._add_user_metrics_sheet(wb, start_date, end_date)
        
        # Booking metrics sheet
        await self._add_booking_metrics_sheet(wb, start_date, end_date)
        
        # Payment metrics sheet
        await self._add_payment_metrics_sheet(wb, start_date, end_date)
        
        # Trip metrics sheet
        await self._add_trip_metrics_sheet(wb, start_date, end_date)
        
        wb.save(file_path)

    async def _add_user_metrics_sheet(self, wb: Workbook, start_date: datetime, end_date: datetime) -> None:
        """Add user metrics sheet to workbook."""
        ws = wb.create_sheet("Users")
        query = """
        SELECT date, new_registrations, kyc_verified_count, driver_upgrades
        FROM openride_analytics.mv_daily_user_metrics
        WHERE date >= {start_date:Date} AND date <= {end_date:Date}
        ORDER BY date
        """
        result = self.ch.query(query, parameters={"start_date": start_date.date(), "end_date": end_date.date()})
        
        ws.append(["Date", "New Registrations", "KYC Verified", "Driver Upgrades"])
        for row in result.result_rows:
            ws.append(row)

    async def _add_booking_metrics_sheet(self, wb: Workbook, start_date: datetime, end_date: datetime) -> None:
        """Add booking metrics sheet to workbook."""
        ws = wb.create_sheet("Bookings")
        query = """
        SELECT date, bookings_created, bookings_confirmed, bookings_cancelled, total_booking_value
        FROM openride_analytics.mv_daily_booking_metrics
        WHERE date >= {start_date:Date} AND date <= {end_date:Date}
        ORDER BY date
        """
        result = self.ch.query(query, parameters={"start_date": start_date.date(), "end_date": end_date.date()})
        
        ws.append(["Date", "Created", "Confirmed", "Cancelled", "Total Value"])
        for row in result.result_rows:
            ws.append(row)

    async def _add_payment_metrics_sheet(self, wb: Workbook, start_date: datetime, end_date: datetime) -> None:
        """Add payment metrics sheet to workbook."""
        ws = wb.create_sheet("Payments")
        query = """
        SELECT date, payments_initiated, payments_success, payments_failed, total_revenue
        FROM openride_analytics.mv_daily_payment_metrics
        WHERE date >= {start_date:Date} AND date <= {end_date:Date}
        GROUP BY date
        ORDER BY date
        """
        result = self.ch.query(query, parameters={"start_date": start_date.date(), "end_date": end_date.date()})
        
        ws.append(["Date", "Initiated", "Success", "Failed", "Revenue"])
        for row in result.result_rows:
            ws.append(row)

    async def _add_trip_metrics_sheet(self, wb: Workbook, start_date: datetime, end_date: datetime) -> None:
        """Add trip metrics sheet to workbook."""
        ws = wb.create_sheet("Trips")
        query = """
        SELECT date, trips_started, trips_completed, trips_cancelled, total_driver_earnings
        FROM openride_analytics.mv_daily_trip_metrics
        WHERE date >= {start_date:Date} AND date <= {end_date:Date}
        ORDER BY date
        """
        result = self.ch.query(query, parameters={"start_date": start_date.date(), "end_date": end_date.date()})
        
        ws.append(["Date", "Started", "Completed", "Cancelled", "Driver Earnings"])
        for row in result.result_rows:
            ws.append(row)
